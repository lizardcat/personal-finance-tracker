import csv
import json
import os
from datetime import datetime, date
from decimal import Decimal
from models import db, Transaction, BudgetCategory, Milestone, User
from utils import ensure_directory_exists, format_currency
import logging

logger = logging.getLogger(__name__)

class ExportService:
    """Service for exporting financial data to various formats"""
    
    def __init__(self):
        self.export_dir = ensure_directory_exists('exports')
    
    def export_transactions(self, user_id, format='csv', start_date=None, end_date=None, category_id=None):
        """Export transactions to specified format"""
        # Build query
        query = Transaction.query.filter_by(user_id=user_id)
        
        if start_date:
            query = query.filter(Transaction.transaction_date >= start_date)
        if end_date:
            query = query.filter(Transaction.transaction_date <= end_date)
        if category_id:
            query = query.filter(Transaction.category_id == category_id)
        
        transactions = query.order_by(Transaction.transaction_date.desc()).all()
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        user = User.query.get(user_id)
        filename = f"transactions_{user.username}_{timestamp}.{format}"
        filepath = os.path.join(self.export_dir, filename)
        
        if format == 'csv':
            return self._export_transactions_csv(transactions, filepath)
        elif format == 'json':
            return self._export_transactions_json(transactions, filepath)
        elif format == 'xlsx':
            return self._export_transactions_xlsx(transactions, filepath)
        else:
            raise ValueError(f"Unsupported export format: {format}")
    
    def export_budget_summary(self, user_id, format='csv', month=None, year=None):
        """Export budget summary to specified format"""
        if not month:
            month = date.today().month
        if not year:
            year = date.today().year
        
        categories = BudgetCategory.query.filter_by(user_id=user_id).all()
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        user = User.query.get(user_id)
        filename = f"budget_summary_{user.username}_{year}_{month:02d}_{timestamp}.{format}"
        filepath = os.path.join(self.export_dir, filename)
        
        if format == 'csv':
            return self._export_budget_csv(categories, filepath, user_id, month, year)
        elif format == 'json':
            return self._export_budget_json(categories, filepath, user_id, month, year)
        else:
            raise ValueError(f"Unsupported export format: {format}")
    
    def export_milestones(self, user_id, format='csv', status=None):
        """Export milestones to specified format"""
        query = Milestone.query.filter_by(user_id=user_id)
        
        if status == 'completed':
            query = query.filter_by(completed=True)
        elif status == 'active':
            query = query.filter_by(completed=False)
        
        milestones = query.order_by(Milestone.target_date.asc().nullslast()).all()
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        user = User.query.get(user_id)
        filename = f"milestones_{user.username}_{timestamp}.{format}"
        filepath = os.path.join(self.export_dir, filename)
        
        if format == 'csv':
            return self._export_milestones_csv(milestones, filepath)
        elif format == 'json':
            return self._export_milestones_json(milestones, filepath)
        else:
            raise ValueError(f"Unsupported export format: {format}")
    
    def export_full_backup(self, user_id):
        """Export complete user data backup"""
        user = User.query.get(user_id)
        if not user:
            raise ValueError("User not found")
        
        # Collect all user data
        backup_data = {
            'user': {
                'username': user.username,
                'email': user.email,
                'default_currency': user.default_currency,
                'monthly_income': float(user.monthly_income) if user.monthly_income else 0,
                'created_at': user.created_at.isoformat() if user.created_at else None
            },
            'budget_categories': [],
            'transactions': [],
            'milestones': [],
            'export_date': datetime.now().isoformat(),
            'version': '1.0'
        }
        
        # Add budget categories
        for category in user.budget_categories:
            backup_data['budget_categories'].append({
                'name': category.name,
                'allocated_amount': float(category.allocated_amount),
                'available_amount': float(category.available_amount),
                'category_type': category.category_type,
                'color': category.color,
                'created_at': category.created_at.isoformat() if category.created_at else None
            })
        
        # Add transactions
        for transaction in user.transactions:
            backup_data['transactions'].append({
                'amount': float(transaction.amount),
                'currency': transaction.currency,
                'description': transaction.description,
                'transaction_type': transaction.transaction_type,
                'transaction_date': transaction.transaction_date.isoformat() if transaction.transaction_date else None,
                'payee': transaction.payee,
                'account': transaction.account,
                'tags': transaction.tags,
                'category_name': transaction.budget_category.name if transaction.budget_category else None,
                'created_at': transaction.created_at.isoformat() if transaction.created_at else None
            })
        
        # Add milestones
        for milestone in user.milestones:
            backup_data['milestones'].append({
                'name': milestone.name,
                'description': milestone.description,
                'target_amount': float(milestone.target_amount),
                'current_amount': float(milestone.current_amount),
                'target_date': milestone.target_date.isoformat() if milestone.target_date else None,
                'completed': milestone.completed,
                'completed_date': milestone.completed_date.isoformat() if milestone.completed_date else None,
                'category': milestone.category,
                'created_at': milestone.created_at.isoformat() if milestone.created_at else None
            })
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"backup_{user.username}_{timestamp}.json"
        filepath = os.path.join(self.export_dir, filename)
        
        # Write backup
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(backup_data, f, indent=2, ensure_ascii=False)
        
        logger.info(f"Full backup created for user {user.username}: {filename}")
        
        return {
            'filepath': filepath,
            'filename': filename,
            'size_bytes': os.path.getsize(filepath),
            'items_count': {
                'categories': len(backup_data['budget_categories']),
                'transactions': len(backup_data['transactions']),
                'milestones': len(backup_data['milestones'])
            }
        }
    
    def _export_transactions_csv(self, transactions, filepath):
        """Export transactions to CSV format"""
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Write header
            writer.writerow([
                'Date', 'Description', 'Category', 'Type', 'Amount', 'Currency',
                'Payee', 'Account', 'Tags', 'Created At'
            ])
            
            # Write data
            for transaction in transactions:
                writer.writerow([
                    transaction.transaction_date.isoformat() if transaction.transaction_date else '',
                    transaction.description,
                    transaction.budget_category.name if transaction.budget_category else '',
                    transaction.transaction_type,
                    float(transaction.amount),
                    transaction.currency,
                    transaction.payee or '',
                    transaction.account,
                    transaction.tags or '',
                    transaction.created_at.isoformat() if transaction.created_at else ''
                ])
        
        logger.info(f"Exported {len(transactions)} transactions to CSV: {os.path.basename(filepath)}")
        
        return {
            'filepath': filepath,
            'filename': os.path.basename(filepath),
            'format': 'csv',
            'records_count': len(transactions),
            'size_bytes': os.path.getsize(filepath)
        }
    
    def _export_transactions_json(self, transactions, filepath):
        """Export transactions to JSON format"""
        data = {
            'export_date': datetime.now().isoformat(),
            'total_records': len(transactions),
            'transactions': []
        }
        
        for transaction in transactions:
            data['transactions'].append({
                'id': transaction.id,
                'date': transaction.transaction_date.isoformat() if transaction.transaction_date else None,
                'description': transaction.description,
                'category': transaction.budget_category.name if transaction.budget_category else None,
                'category_color': transaction.budget_category.color if transaction.budget_category else None,
                'type': transaction.transaction_type,
                'amount': float(transaction.amount),
                'currency': transaction.currency,
                'payee': transaction.payee,
                'account': transaction.account,
                'tags': transaction.tags,
                'recurring': transaction.recurring,
                'recurring_period': transaction.recurring_period,
                'created_at': transaction.created_at.isoformat() if transaction.created_at else None
            })
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        logger.info(f"Exported {len(transactions)} transactions to JSON: {os.path.basename(filepath)}")
        
        return {
            'filepath': filepath,
            'filename': os.path.basename(filepath),
            'format': 'json',
            'records_count': len(transactions),
            'size_bytes': os.path.getsize(filepath)
        }
    
    def _export_transactions_xlsx(self, transactions, filepath):
        """Export transactions to Excel format"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError("openpyxl is required for Excel export")
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Transactions"
        
        # Headers
        headers = [
            'Date', 'Description', 'Category', 'Type', 'Amount', 'Currency',
            'Payee', 'Account', 'Tags', 'Created At'
        ]
        
        # Style headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Data rows
        for row, transaction in enumerate(transactions, 2):
            sheet.cell(row=row, column=1, value=transaction.transaction_date)
            sheet.cell(row=row, column=2, value=transaction.description)
            sheet.cell(row=row, column=3, value=transaction.budget_category.name if transaction.budget_category else '')
            sheet.cell(row=row, column=4, value=transaction.transaction_type)
            sheet.cell(row=row, column=5, value=float(transaction.amount))
            sheet.cell(row=row, column=6, value=transaction.currency)
            sheet.cell(row=row, column=7, value=transaction.payee or '')
            sheet.cell(row=row, column=8, value=transaction.account)
            sheet.cell(row=row, column=9, value=transaction.tags or '')
            sheet.cell(row=row, column=10, value=transaction.created_at)
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(filepath)
        
        logger.info(f"Exported {len(transactions)} transactions to Excel: {os.path.basename(filepath)}")
        
        return {
            'filepath': filepath,
            'filename': os.path.basename(filepath),
            'format': 'xlsx',
            'records_count': len(transactions),
            'size_bytes': os.path.getsize(filepath)
        }
    
    def _export_budget_csv(self, categories, filepath, user_id, month, year):
        """Export budget summary to CSV format"""
        from utils import get_month_range
        from sqlalchemy import func
        
        start_date, end_date = get_month_range(year, month)
        
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Write header
            writer.writerow([
                'Category', 'Type', 'Allocated Amount', 'Actual Amount', 'Available Amount',
                'Percentage Used', 'Color', 'Status'
            ])
            
            # Write data
            for category in categories:
                # Calculate actual spending for the period
                actual_amount = db.session.query(func.sum(Transaction.amount))\
                    .filter_by(user_id=user_id, category_id=category.id)\
                    .filter(Transaction.transaction_date >= start_date)\
                    .filter(Transaction.transaction_date <= end_date)
                
                if category.category_type == 'expense':
                    actual_amount = actual_amount.filter_by(transaction_type='expense').scalar() or Decimal('0')
                elif category.category_type == 'income':
                    actual_amount = actual_amount.filter_by(transaction_type='income').scalar() or Decimal('0')
                else:
                    actual_amount = actual_amount.filter_by(transaction_type='transfer').scalar() or Decimal('0')
                
                available_amount = category.allocated_amount - actual_amount
                percentage_used = (float(actual_amount) / float(category.allocated_amount) * 100) if category.allocated_amount > 0 else 0
                
                # Determine status
                status = 'Good'
                if category.category_type == 'expense':
                    if percentage_used > 100:
                        status = 'Over Budget'
                    elif percentage_used > 90:
                        status = 'Warning'
                
                writer.writerow([
                    category.name,
                    category.category_type,
                    float(category.allocated_amount),
                    float(actual_amount),
                    float(available_amount),
                    round(percentage_used, 2),
                    category.color,
                    status
                ])
        
        logger.info(f"Exported budget summary to CSV: {os.path.basename(filepath)}")
        
        return {
            'filepath': filepath,
            'filename': os.path.basename(filepath),
            'format': 'csv',
            'records_count': len(categories),
            'size_bytes': os.path.getsize(filepath)
        }
    
    def _export_budget_json(self, categories, filepath, user_id, month, year):
        """Export budget summary to JSON format"""
        from utils import get_month_range
        from sqlalchemy import func
        
        start_date, end_date = get_month_range(year, month)
        
        data = {
            'export_date': datetime.now().isoformat(),
            'period': {
                'month': month,
                'year': year,
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat()
            },
            'categories': []
        }
        
        for category in categories:
            # Calculate actual spending for the period
            actual_amount = db.session.query(func.sum(Transaction.amount))\
                .filter_by(user_id=user_id, category_id=category.id)\
                .filter(Transaction.transaction_date >= start_date)\
                .filter(Transaction.transaction_date <= end_date)
            
            if category.category_type == 'expense':
                actual_amount = actual_amount.filter_by(transaction_type='expense').scalar() or Decimal('0')
            elif category.category_type == 'income':
                actual_amount = actual_amount.filter_by(transaction_type='income').scalar() or Decimal('0')
            else:
                actual_amount = actual_amount.filter_by(transaction_type='transfer').scalar() or Decimal('0')
            
            data['categories'].append({
                'name': category.name,
                'type': category.category_type,
                'allocated_amount': float(category.allocated_amount),
                'actual_amount': float(actual_amount),
                'available_amount': float(category.allocated_amount - actual_amount),
                'percentage_used': (float(actual_amount) / float(category.allocated_amount) * 100) if category.allocated_amount > 0 else 0,
                'color': category.color,
                'created_at': category.created_at.isoformat() if category.created_at else None
            })
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        logger.info(f"Exported budget summary to JSON: {os.path.basename(filepath)}")
        
        return {
            'filepath': filepath,
            'filename': os.path.basename(filepath),
            'format': 'json',
            'records_count': len(categories),
            'size_bytes': os.path.getsize(filepath)
        }
    
    def _export_milestones_csv(self, milestones, filepath):
        """Export milestones to CSV format"""
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Write header
            writer.writerow([
                'Name', 'Description', 'Category', 'Target Amount', 'Current Amount',
                'Progress %', 'Target Date', 'Completed', 'Completed Date', 'Created At'
            ])
            
            # Write data
            for milestone in milestones:
                writer.writerow([
                    milestone.name,
                    milestone.description or '',
                    milestone.category,
                    float(milestone.target_amount),
                    float(milestone.current_amount),
                    milestone.progress_percentage,
                    milestone.target_date.isoformat() if milestone.target_date else '',
                    'Yes' if milestone.completed else 'No',
                    milestone.completed_date.isoformat() if milestone.completed_date else '',
                    milestone.created_at.isoformat() if milestone.created_at else ''
                ])
        
        logger.info(f"Exported {len(milestones)} milestones to CSV: {os.path.basename(filepath)}")
        
        return {
            'filepath': filepath,
            'filename': os.path.basename(filepath),
            'format': 'csv',
            'records_count': len(milestones),
            'size_bytes': os.path.getsize(filepath)
        }
    
    def _export_milestones_json(self, milestones, filepath):
        """Export milestones to JSON format"""
        data = {
            'export_date': datetime.now().isoformat(),
            'total_records': len(milestones),
            'milestones': []
        }
        
        for milestone in milestones:
            data['milestones'].append({
                'id': milestone.id,
                'name': milestone.name,
                'description': milestone.description,
                'category': milestone.category,
                'target_amount': float(milestone.target_amount),
                'current_amount': float(milestone.current_amount),
                'progress_percentage': milestone.progress_percentage,
                'target_date': milestone.target_date.isoformat() if milestone.target_date else None,
                'completed': milestone.completed,
                'completed_date': milestone.completed_date.isoformat() if milestone.completed_date else None,
                'is_overdue': milestone.is_overdue,
                'created_at': milestone.created_at.isoformat() if milestone.created_at else None
            })
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        logger.info(f"Exported {len(milestones)} milestones to JSON: {os.path.basename(filepath)}")
        
        return {
            'filepath': filepath,
            'filename': os.path.basename(filepath),
            'format': 'json',
            'records_count': len(milestones),
            'size_bytes': os.path.getsize(filepath)
        }

# Create singleton instance
export_service = ExportService()